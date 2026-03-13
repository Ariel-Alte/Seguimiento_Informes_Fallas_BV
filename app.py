"""
PISE — Extractor de Informes de Inspección de Seguridad Estática
Versión: 3.0
Plataforma: Gradio + Hugging Face Spaces

Flujo:
  1. Extrae encabezado (vehículo, línea, informe N°, fecha, código PISE)
  2. Recorre secciones 1-10, extrae todos los valores por ítem
  3. Lee sección D. Conclusiones → clasifica cada ítem observado
  4. Cruza por ítem → columna "Clasificación"
  5. Exporta Excel con 3 hojas: todos los datos, observaciones, solo X
"""

import gradio as gr
import pdfplumber
import pandas as pd
import re
import io
import unicodedata
from pathlib import Path


# ══════════════════════════════════════════════════════════
#  UTILIDADES
# ══════════════════════════════════════════════════════════

def norm(texto: str) -> str:
    """Elimina tildes, normaliza espacios, mayúsculas."""
    if not texto:
        return ""
    t = unicodedata.normalize("NFD", str(texto))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t).strip().upper()


def limpia(val) -> str:
    """Limpia una celda de tabla."""
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val)).strip()


def es_item(texto: str) -> bool:
    """Verifica si el texto es un número de ítem válido (1.1, 2.15.1, etc.)."""
    return bool(re.match(r"^\d+(?:\.\d+)*$", limpia(texto)))


def extraer_rango(descripcion: str) -> str:
    """Extrae el valor/rango esperado embebido en la descripción."""
    d = norm(descripcion)
    patrones = [
        r"[\d.,]+\s*(?:\+/-|±|\+\s*/\s*-)\s*[\d.,]+\s*\w*",
        r"[≥≤<>]\s*[\d.,]+\s*(?:MM|KG|KPA|CM|%|MM\.)?",
        r"\([\d.,]+\s*[≤≥<>]\s*\w+\s*[≤≥<>]\s*[\d.,]+\)",
        r"(?:ENTRE)\s+[\d.,]+\s*(?:Y)\s*[\d.,]+",
        r"MAX(?:IMO)?\s*[:\s]*[\d.,]+",
        r"MIN(?:IMO)?\s*[:\s]*[\d.,]+",
        r"[\d.,]+\s*(?:MM|KPA|KG/CM2|CMHG)\b",
    ]
    for p in patrones:
        m = re.search(p, d)
        if m:
            return m.group(0).strip()
    return ""


# ══════════════════════════════════════════════════════════
#  EXTRACCIÓN DE PÁGINAS
# ══════════════════════════════════════════════════════════

def leer_pdf(pdf_path: str) -> list:
    """Lee todas las páginas del PDF, devuelve texto y tablas por página."""
    paginas = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            texto = page.extract_text() or ""
            tablas = page.extract_tables() or []
            paginas.append({"num": i + 1, "texto": texto, "tablas": tablas})
    return paginas


# ══════════════════════════════════════════════════════════
#  ENCABEZADO
# ══════════════════════════════════════════════════════════

def extraer_encabezado(paginas: list) -> dict:
    """Extrae metadatos del informe desde las primeras páginas."""
    enc = {
        "vehiculo": "", "linea": "", "lugar": "", "marca": "",
        "informe_nro": "", "inspeccion_nro": "", "fecha_inspeccion": "",
        "pise_codigo": "", "codigo_pise_informe": "", "_filename": "",
    }
    texto = "\n".join(p["texto"] for p in paginas[:4])

    m = re.search(r"(PISE-[A-Z]+-\d+(?::\s*REV\.\d+)?)", texto, re.IGNORECASE)
    if m:
        enc["pise_codigo"] = m.group(1).strip()

    m = re.search(r"Informe\s*N[°º\.][:\s]*(\d+)", texto, re.IGNORECASE)
    if m:
        enc["informe_nro"] = m.group(1)

    m = re.search(r"Inspecci[oó]n\s*N[°º\.][:\s]*(\d+)", texto, re.IGNORECASE)
    if m:
        enc["inspeccion_nro"] = m.group(1)

    m = re.search(r"FECHA\s+DE\s+INSP[\.:\s]+(\d{1,2}/\d{1,2}/\d{2,4})", texto, re.IGNORECASE)
    if m:
        enc["fecha_inspeccion"] = m.group(1)
    else:
        m = re.search(r"Fecha[:\s]+(\d{1,2}/\d{1,2}/\d{2,4})", texto, re.IGNORECASE)
        if m:
            enc["fecha_inspeccion"] = m.group(1)

    m = re.search(r"VEHICULO\s*[:\s]+([A-Z]+\s*\d+)", texto, re.IGNORECASE)
    if m:
        enc["vehiculo"] = m.group(1).strip()
    else:
        m = re.search(r"LOCOMOTORA\s+N[°º][:\s]+([A-Z]\s*\d+)", texto, re.IGNORECASE)
        if m:
            enc["vehiculo"] = m.group(1).replace(" ", "")
        else:
            m = re.search(r"IDENTIFICACI[OÓ]N\s*[:\s]+([A-Z0-9\s]+?)[\n\r]", texto, re.IGNORECASE)
            if m:
                enc["vehiculo"] = m.group(1).strip()

    m = re.search(r"L[ÍI]NEA\s*[:\s]+([A-ZÁÉÍÓÚ\s]+?)[\n\r]", texto, re.IGNORECASE)
    if m:
        enc["linea"] = m.group(1).strip()
    m = re.search(r"LUGAR\s*[:\s]+([A-ZÁÉÍÓÚ\s]+?)[\n\r]", texto, re.IGNORECASE)
    if m:
        enc["lugar"] = m.group(1).strip()
    m = re.search(r"MARCA\s*[:\s]+([A-ZÁÉÍÓÚ\s]+?)[\n\r]", texto, re.IGNORECASE)
    if m:
        enc["marca"] = m.group(1).strip()

    if enc["pise_codigo"] and enc["informe_nro"]:
        enc["codigo_pise_informe"] = f"{enc['pise_codigo']}/{enc['informe_nro']}"

    return enc


# ══════════════════════════════════════════════════════════
#  D. CONCLUSIONES — CLASIFICACIÓN DE ÍTEMS
# ══════════════════════════════════════════════════════════

def _limpiar_texto_conc(texto: str) -> str:
    """Elimina caracteres de control del texto de Conclusiones."""
    texto = "".join(c for c in texto if unicodedata.category(c)[0] != "C")
    return re.sub(r"\s+", " ", texto).strip()


def extraer_clasificaciones(paginas: list) -> dict:
    """
    Lee D. Conclusiones usando regex directo (igual al código original).
    Retorna {item: (categoria, item_par, descripcion)}.

    Patrón: Ítems X // Y (descripción)  o  Ítem X (descripción)
    Ambos ítems del par quedan en el dict apuntando al otro.
    """
    clasificaciones = {}

    # Buscar la página real de Conclusiones (no el índice)
    texto_conc = ""
    for pag in paginas:
        t = pag["texto"]
        if (re.search(r"[D]\.\s*CONCLUSIONES?", t, re.IGNORECASE) and
                re.search(r"normales?\s*\(?no\s+cr[ií]tic", t, re.IGNORECASE)):
            texto_conc = _limpiar_texto_conc(t)
            break

    if not texto_conc:
        return clasificaciones

    # Encabezados de categoría en el PDF
    ENCABEZADOS = [
        ("REPARADA",  r"reparad[ao]s?\s+durante\s+la\s+inspecci[oó]n\s*:"),
        ("NORMAL",    r"normales?\s*\(?no\s+cr[ií]tic[ao]s?\)?\s*:"),
        ("CRITICA",   r"cr[ií]tic[ao]s?\s*:"),
        ("RECHAZADA", r"rechazad[ao]s?\s*:"),
    ]

    posiciones = []
    for cat, patron in ENCABEZADOS:
        m = re.search(patron, texto_conc, re.IGNORECASE)
        if m:
            posiciones.append((m.start(), m.end(), cat))
    posiciones.sort()

    # Regex unificado — 4 formatos de descripción sin paréntesis obligatorio:
    #   (texto)  |  : texto  |  - texto  |  Texto sin delimitador (mayúscula, ≥5 chars)
    # Prefijo "Ítems" opcional para cubrir líneas sin él.
    PATRON_ITEM = (
        r"(?:(?:[ÍIíi]tems?|items?)\s+)?"          # prefijo opcional
        r"(\d+(?:\s*\.\s*\d+)+)"                # ítem 1
        r"(?:\s*//\s*(\d+(?:\s*\.\s*\d+)+))?" # // ítem 2 (opcional)
        r"\s*"
        r"(?:"
            r"\(([^()]+)\)"                         # grupo 3: (descripción)
            r"|:\s*([^.;\n]+)"                      # grupo 4: : descripción
            r"|-\s*([^.;\n]+)"                      # grupo 5: - descripción
            r"|([A-ZÁÉÍÓÚÑ][^.;\n(]{4,})"           # grupo 6: sin delimitador
        r")"
    )

    def _limpiar_desc(desc: str) -> str:
        return desc.strip().rstrip(".,) ").strip()

    for i, (_, fin_header, cat) in enumerate(posiciones):
        fin_bloque = posiciones[i + 1][0] if i + 1 < len(posiciones) else len(texto_conc)
        bloque = texto_conc[fin_header:fin_bloque]

        if re.search(r"NO\s+SE\s+(?:OBSERVARON|REALIZARON)", norm(bloque)):
            continue

        matches_item = list(re.finditer(PATRON_ITEM, bloque, re.IGNORECASE))

        if matches_item:
            for match in matches_item:
                item1 = re.sub(r"\s+", "", match.group(1))
                item2 = re.sub(r"\s+", "", match.group(2) or "")
                desc  = _limpiar_desc(next((g for g in match.groups()[2:] if g), ""))
                clasificaciones[item1] = (cat, item2, desc)
                if item2:
                    clasificaciones[item2] = (cat, item1, desc)
        elif cat == "REPARADA":
            # Texto libre sin número de ítem (ej: "Se reparó pérdida de combustible.")
            # Nota: _limpiar_texto_conc colapsa a una línea, así que ^ con MULTILINE no funciona.
            # Se usa el marcador ➢ / • como delimitador de frase.
            for m in re.finditer(r"[➢•►\-]\s*([A-Za-záéíóúñÁÉÍÓÚÑ][^•➢►.]{5,})\.",
                                  bloque, re.IGNORECASE):
                desc = _limpiar_desc(m.group(1))
                clasificaciones.setdefault("_REPARADA_LIBRE", (cat, "", desc))

    return clasificaciones


def extraer_obs_texto(texto_pagina: str) -> dict:
    """
    Extrae observaciones del bloque OBSERVACIONES: al pie de cada hoja.
    Retorna {item: texto_observacion}.
    """
    obs = {}
    m = re.search(
        r"OBSERVACIONES\s*:?\s*\n(.*?)(?=REFERENCIAS|FIRMA\s+INSPECTOR|$)",
        texto_pagina, re.DOTALL | re.IGNORECASE
    )
    if not m:
        return obs
    bloque = m.group(1)
    for linea in bloque.split("\n"):
        linea = linea.strip()
        if not linea:
            continue
        mo = re.match(r"^(\d+(?:\.\d+)*)[.\s]+(.+)$", linea)
        if mo:
            texto_obs = mo.group(2).strip()
            # Limpiar prefijos tipo "// 1.18 " que a veces quedan del texto
            texto_obs = re.sub(r"^//\s*\d+(?:\.\d+)*\s*", "", texto_obs).strip()
            obs[mo.group(1)] = texto_obs
    return obs


# ══════════════════════════════════════════════════════════
#  DETECCIÓN DE SECCIÓN
# ══════════════════════════════════════════════════════════

SECCIONES = [
    (r"1[\.\s]+PAR(?:ES)?\s+MONTADOS?[\.\s]+RUEDAS?",                   "1_RUEDAS"),
    (r"2a[\.\s]+BOGIE\s*1",                                              "2A_BOGIE1"),
    (r"2b[\.\s]+BOGIE\s*2",                                              "2B_BOGIE2"),
    (r"2[\.\s]+BOGIE\s*1\s*[–\-]\s*ESTRUCTURA",                         "2A_BOGIE1"),
    (r"2[\.\s]+BOGIE\s*2\s*[–\-]\s*ESTRUCTURA",                         "2B_BOGIE2"),
    # PISE-024: "2. BOGIES 1." y "2. BOGIE 2." con punto final
    (r"2[\.\s]+BOGIES?\s+1\b",                                           "2A_BOGIE1"),
    (r"2[\.\s]+BOGIES?\s+2\b",                                           "2B_BOGIE2"),
    (r"2c[\.\s]+MESA\s+CENTRAL",                                         "2C_MESA_CENTRAL"),
    (r"(?m)^2[\.\s]+BOGIES?\b",                                          "2_BOGIES"),
    (r"3[\.\s a-z]+FRENOS?.*MEC[AÁ]NICA",                               "3A_FRENOS_MECANICA"),
    (r"3[\.\s b]+FRENOS?.*NEUM[AÁ]TICA",                                "3B_FRENOS_NEUMATICA"),
    # PISE-024: "3. SISTEMA DE FRENOS." — frenos mecánica genérico
    (r"3[\.\s]+SISTEMA\s+DE\s+FRENOS?\b",                               "3A_FRENOS_MECANICA"),
    (r"4[\.\s]+.*TRACCI[OÓ]N.*PUNTA\s+CABINA",                          "4A_TRACCION_CABINA"),
    (r"4[\.\s]+.*TRACCI[OÓ]N.*PUNTA\s+CONTRARIA",                       "4B_TRACCION_CONTRARIA"),
    (r"4[\.\s]+SISTEMAS?\s+DE\s+TRACCI[OÓ]N\s+Y\s+CHOQUE",             "4_TRACCION_CHOQUE"),
    (r"5[\.\s]+ELEMENTOS?\s+BAJO\s+BASTIDOR",                            "5_BAJO_BASTIDOR"),
    (r"6[\.\s]+INTERIOR\s+(?:DE\s+)?SAL[OÓ]N",                         "6_INTERIOR_SALON"),
    (r"6[\.\s]+INTERIOR\s+CABINA\s+DE\s+CONDUCCI[OÓ]N",                "6_INTERIOR_CABINA"),
    (r"7[\.\s]+ESTRUCTURA\s+EXTERIOR",                                   "7_ESTRUCTURA_EXT"),
    (r"8[\.\s]+MOTORES?\s+DI[EÉ]SEL\s+Y\s+GENERADORES?",               "8_MOTORES_DIESEL"),
    (r"8[\.\s]+SALA\s+DE\s+M[AÁ]QUINAS",                               "8_SALA_MAQUINAS"),
    # PISE-024: "8. FURGÓN Y MOTOGENERADOR."
    (r"8[\.\s]+FURG[OÓ]N\s+Y\s+MOTOGENERADOR",                         "8_FURGON_MOTOGEN"),
    (r"9[\.\s]+INTERIOR\s+CABINA\s+DE\s+CONDUCCI[OÓ]N",                "9_INTERIOR_CABINA"),
    (r"9[\.\s]+EQUIPAMIENTO\s+SOBRE\s+TECHO",                           "9_TECHO"),
    (r"10[\.\s]+PRUEBA\s+EST[AÁ]TICA\s+DEL?\s+ATS",                    "10_ATS"),
]


def detectar_seccion(texto: str):
    t = norm(texto)
    for patron, nombre in SECCIONES:
        if re.search(patron, t, re.IGNORECASE | re.MULTILINE):
            return nombre
    return None


# ══════════════════════════════════════════════════════════
#  HELPERS DE TABLAS
# ══════════════════════════════════════════════════════════

def buscar_contexto_col(tabla, header_idx: int, col_idx: int) -> str:
    """Busca BOGIE/EJE en filas anteriores para una columna dada."""
    for i in range(header_idx - 1, -1, -1):
        fila = tabla[i]
        for delta in range(min(5, len(fila))):
            for idx in [col_idx - delta, col_idx + delta]:
                if 0 <= idx < len(fila):
                    v = norm(limpia(fila[idx]))
                    if re.search(r"(BOGIE|EJE)\s*(?:N[°º])?\s*\d+", v):
                        return limpia(fila[idx])
    return ""


def detectar_header_ruedas(tabla):
    """
    Detecta la fila de encabezados con RUEDA N.
    Retorna (header_idx, col_map) donde
    col_map = {col_idx: (bogie, rueda, lado, contexto)}.
    """
    for i, fila in enumerate(tabla[:7]):
        contenido = norm(" ".join(limpia(c) for c in fila if c))
        if re.search(r"RUEDA\s*\d+", contenido):
            col_map = {}
            for j, celda in enumerate(fila):
                v = limpia(celda)
                vn = norm(v)
                rm = re.search(r"RUEDA\s*(\d+)", vn)
                if rm:
                    rueda = rm.group(1)
                    lado_m = re.search(r"\(([ID])\)", v)
                    lado = lado_m.group(1) if lado_m else ""
                    ctx = buscar_contexto_col(tabla, i, j)
                    bg_m = re.search(r"BOGIE\s*(?:N[°º])?\s*(\d+)", norm(ctx))
                    bogie = bg_m.group(1) if bg_m else ""
                    col_map[j] = (bogie, rueda, lado, ctx)
            return i, col_map
    return None, {}


# ══════════════════════════════════════════════════════════
#  META Y FILA BASE
# ══════════════════════════════════════════════════════════

def meta(enc: dict) -> dict:
    return {
        "Archivo PDF":          enc.get("_filename", ""),
        "Código PISE/Informe":  enc.get("codigo_pise_informe", ""),
        "Vehículo":             enc.get("vehiculo", ""),
        "Línea":                enc.get("linea", ""),
        "Lugar":                enc.get("lugar", ""),
        "Fecha inspección":     enc.get("fecha_inspeccion", ""),
        "Informe N°":           enc.get("informe_nro", ""),
        "PISE código":          enc.get("pise_codigo", ""),
    }


def fila_base(enc, seccion, item, desc, bogie="", rueda="", lado="",
              ubicacion="", v_esp="", v_med="", obs="") -> dict:
    return {
        **meta(enc),
        "Sección":        seccion,
        "Ítem técnico":   item,
        "Descripción":    desc,
        "Bogie":          bogie,
        "Rueda":          rueda,
        "Lado":           lado,
        "Ubicación":      ubicacion,
        "Valor esperado": v_esp,
        "Valor medido":   v_med,
        "Observación":    obs,
        "Ítem técnico 2": "",   # par del ítem (si viene de Conclusiones)
        "Clasificación":  "",   # se rellena después con Conclusiones
    }


# ══════════════════════════════════════════════════════════
#  PARSERS POR SECCIÓN
# ══════════════════════════════════════════════════════════

def _detectar_todos_headers_ruedas(tabla):
    """
    Devuelve lista de (header_idx, col_map) para todos los bloques RUEDA en la tabla.
    En PISE-024 una misma tabla tiene BOGIE 1 y BOGIE 2 como sub-bloques separados.
    """
    bloques = []
    i = 0
    while i < len(tabla):
        fila = tabla[i]
        contenido = norm(" ".join(limpia(c) for c in fila if c))
        if re.search(r"RUEDA\s*\d+", contenido):
            col_map = {}
            for j, celda in enumerate(fila):
                v = limpia(celda)
                vn = norm(v)
                rm = re.search(r"RUEDA\s*(\d+)", vn)
                if rm:
                    rueda = rm.group(1)
                    lado_m = re.search(r"\(([ID])\)", v)
                    lado = lado_m.group(1) if lado_m else ""
                    ctx = buscar_contexto_col(tabla, i, j)
                    bg_m = re.search(r"BOGIE\s*(?:N[°º])?\s*(\d+)", norm(ctx))
                    bogie = bg_m.group(1) if bg_m else ""
                    # Buscar BOGIE en fila 0 del bloque (la fila con "ITEMS" o "BOGIE N")
                    if not bogie:
                        for fi_idx in range(i - 1, max(i - 5, -1), -1):
                            fi = tabla[fi_idx]
                            for delta in range(-6, 7):
                                idx = j + delta
                                if 0 <= idx < len(fi):
                                    bm = re.search(r"BOGIE\s*(?:N[°º])?\s*(\d+)",
                                                   norm(limpia(fi[idx])))
                                    if bm:
                                        bogie = bm.group(1)
                                        break
                            if bogie:
                                break
                    col_map[j] = (bogie, rueda, lado, ctx)
            if col_map:
                bloques.append((i, col_map))
        i += 1
    return bloques


def parse_ruedas(pagina: dict, enc: dict, seccion: str) -> list:
    """Extrae TODOS los valores numéricos. Una fila por (ítem, rueda).
    Soporta:
    - Layout estándar: una fila de headers RUEDA por tabla
    - Layout multi-bloque PISE-024: BOGIE 1 y BOGIE 2 en sub-bloques de la misma tabla
    - Layout corrido: datos en col-1 respecto al header RUEDA
    """
    filas = []
    obs_txt = extraer_obs_texto(pagina["texto"])
    for tabla in pagina["tablas"]:
        if not tabla or len(tabla) < 3:
            continue
        bloques = _detectar_todos_headers_ruedas(tabla)
        if not bloques:
            continue
        for b_idx, (header_idx, col_map) in enumerate(bloques):
            col_map = _col_map_corregido(col_map, tabla, header_idx)
            # Filas de datos: desde header+1 hasta el siguiente bloque (o fin)
            fin_bloque = bloques[b_idx + 1][0] if b_idx + 1 < len(bloques) else len(tabla)
            for fila in tabla[header_idx + 1:fin_bloque]:
                item = limpia(fila[0]) if fila else ""
                if not es_item(item):
                    continue
                desc = limpia(fila[1]) if len(fila) > 1 else ""
                v_esp = extraer_rango(desc)
                obs = obs_txt.get(item, "")
                for j, (bogie, rueda, lado, ctx) in col_map.items():
                    if j >= len(fila):
                        continue
                    valor = limpia(fila[j])
                    if not valor or valor in ("-", "N/A"):
                        continue
                    filas.append(fila_base(
                        enc, seccion, item, desc,
                        bogie=bogie, rueda=rueda, lado=lado,
                        ubicacion=ctx, v_esp=v_esp, v_med=valor, obs=obs
                    ))
    return filas


def _col_map_corregido(col_map: dict, tabla: list, header_idx: int) -> dict:
    """
    Detecta si el layout tiene los datos en col-1 respecto al header RUEDA.
    Ocurre en PISE-024 donde pdfplumber desplaza los datos 1 columna a la izquierda.
    Adicionalmente rellena el campo bogie buscando en fila 0 de la tabla.
    """
    if not col_map or header_idx is None:
        return col_map
    for fila in tabla[header_idx + 1:]:
        item = limpia(fila[0]) if fila else ""
        if not es_item(item):
            continue
        cols_j      = [j for j in col_map if j < len(fila) and limpia(fila[j])]
        cols_j_prev = [j for j in col_map if j > 0 and (j-1) < len(fila)
                       and limpia(fila[j-1]) and not limpia(fila[j])]
        if not cols_j and cols_j_prev:
            fila0 = tabla[0] if tabla else []
            nuevo = {}
            for j, (bogie, rueda, lado, ctx) in col_map.items():
                j_new = j - 1
                bogie_real = bogie
                if not bogie_real:
                    for delta in range(-5, 6):
                        idx = j + delta
                        if 0 <= idx < len(fila0):
                            bm = re.search(r"BOGIE\s*(?:N[°º])?\s*(\d+)",
                                           norm(limpia(fila0[idx])))
                            if bm:
                                bogie_real = bm.group(1)
                                break
                nuevo[j_new] = (bogie_real, rueda, lado, ctx)
            return nuevo
        break
    return col_map


def parse_bogies(pagina: dict, enc: dict, seccion: str) -> list:
    """
    X (falla visual) y valores numéricos de huelgos.
    OK puro se omite — no aporta información.
    Soporta dos layouts:
      - Estándar: col_map con RUEDA headers → una fila por rueda
      - PISE-024 columna única: valor en col 2 sin separación por rueda
    """
    filas = []
    obs_txt = extraer_obs_texto(pagina["texto"])
    for tabla in pagina["tablas"]:
        if not tabla or len(tabla) < 3:
            continue
        header_idx, col_map = detectar_header_ruedas(tabla)
        if header_idx is None or not col_map:
            continue

        # Detectar si es layout corrido (PISE-024)
        col_map_usado = _col_map_corregido(col_map, tabla, header_idx)

        # Detectar layout de columna única: todos los datos reales en col 2
        # (bogie completo sin desglose por rueda)
        primera_fila_dato = next(
            (f for f in tabla[header_idx + 1:] if es_item(limpia(f[0]) if f else "")),
            None
        )
        es_col_unica = False
        if primera_fila_dato is not None:
            # Si ninguna col del col_map tiene dato y col 2 tiene dato → col única
            tiene_en_map = any(
                j < len(primera_fila_dato) and limpia(primera_fila_dato[j])
                for j in col_map_usado
            )
            tiene_en_col2 = (len(primera_fila_dato) > 2 and
                             limpia(primera_fila_dato[2]) not in ("", "-", "N/A"))
            if not tiene_en_map and tiene_en_col2:
                es_col_unica = True

        for fila in tabla[header_idx + 1:]:
            item = limpia(fila[0]) if fila else ""
            if not es_item(item):
                continue
            desc = limpia(fila[1]) if len(fila) > 1 else ""
            v_esp = extraer_rango(desc)
            obs = obs_txt.get(item, "")

            if es_col_unica:
                # Layout PISE-024: valores en columnas del col_map_usado (ej cols 2 y 7)
                # que corresponden a EJE1 y EJE2 del bogie. Emitir una fila por columna con dato.
                bogie_ctx = ""
                for fi in tabla[:header_idx]:
                    for c in fi:
                        bm = re.search(r"BOGIE\s*(?:N[°º])?\s*(\d+)", norm(limpia(c)))
                        if bm:
                            bogie_ctx = bm.group(1)
                for j, (bogie, rueda, lado, ctx) in col_map_usado.items():
                    if j >= len(fila):
                        continue
                    valor = limpia(fila[j])
                    vn = norm(valor)
                    if not valor or valor in ("-", "N/A", "OK", "B", "C", "A"):
                        continue
                    if vn == "X" or re.match(r"^[\d.,/]+$", valor.replace(" ", "")):
                        bg = bogie or bogie_ctx
                        filas.append(fila_base(
                            enc, seccion, item, desc,
                            bogie=bg, rueda=rueda, lado=lado,
                            ubicacion=ctx or (f"BOGIE {bg}" if bg else seccion),
                            v_esp=v_esp, v_med=valor, obs=obs
                        ))
            else:
                for j, (bogie, rueda, lado, ctx) in col_map_usado.items():
                    if j >= len(fila):
                        continue
                    valor = limpia(fila[j])
                    vn = norm(valor)
                    if not valor or valor in ("-", "N/A"):
                        continue
                    if vn == "X" or re.match(r"^[\d.,/]+$", valor.replace(" ", "")):
                        filas.append(fila_base(
                            enc, seccion, item, desc,
                            bogie=bogie, rueda=rueda, lado=lado,
                            ubicacion=ctx,
                            v_esp=v_esp, v_med=valor, obs=obs
                        ))
    return filas


def parse_mesa_central(pagina: dict, enc: dict, seccion: str) -> list:
    """Extrae huelgos A-H: A-D → Bogie 1, E-H → Bogie 2."""
    filas = []
    obs_txt = extraer_obs_texto(pagina["texto"])
    LETRAS_B1 = list("ABCD")

    for tabla in pagina["tablas"]:
        if not tabla or len(tabla) < 3:
            continue
        # Detectar fila con 4+ letras A-H
        header_idx = None
        col_letras = []
        for i, fila in enumerate(tabla):
            letras = [(j, limpia(c).strip()) for j, c in enumerate(fila)
                      if limpia(c).strip() in list("ABCDEFGH")]
            if len(letras) >= 4:
                header_idx = i
                col_letras = letras
                break
        if header_idx is None:
            continue
        for fila in tabla[header_idx + 1:]:
            item = limpia(fila[0]) if fila else ""
            if not es_item(item):
                continue
            desc = limpia(fila[1]) if len(fila) > 1 else ""
            v_esp = extraer_rango(desc)
            obs = obs_txt.get(item, "")
            for j, letra in col_letras:
                if j >= len(fila):
                    continue
                valor = limpia(fila[j])
                if not valor or valor in ("-", "N/A", ""):
                    continue
                bogie = "1" if letra in LETRAS_B1 else "2"
                filas.append(fila_base(
                    enc, seccion, item, desc,
                    bogie=bogie, lado=letra,
                    ubicacion=f"BOGIE {bogie} - POS {letra}",
                    v_esp=v_esp, v_med=valor, obs=obs
                ))
    return filas


def parse_frenos_neumatica(pagina: dict, enc: dict, seccion: str) -> list:
    """
    TEÓRICO | REAL | OBS  (locomotora)
    o CONTROL EN CABINA | OBS  (coche).
    Incluye todos los valores medidos.
    """
    filas = []
    obs_txt = extraer_obs_texto(pagina["texto"])
    for tabla in pagina["tablas"]:
        if not tabla or len(tabla) < 3:
            continue
        header_idx = col_teo = col_real = col_obs = None
        for i, fila in enumerate(tabla[:6]):
            cn = norm(" ".join(limpia(c) for c in fila if c))
            if "TEORICO" in cn or "REAL" in cn:
                header_idx = i
                for j, celda in enumerate(fila):
                    c = norm(limpia(celda))
                    if "TEORICO" in c:
                        col_teo = j
                    elif c == "REAL":
                        col_real = j
                    elif "OBSERV" in c:
                        col_obs = j
                break
            elif "CONTROL" in cn and "CABINA" in cn:
                header_idx = i
                for j, celda in enumerate(fila):
                    c = norm(limpia(celda))
                    if "CONTROL" in c and col_real is None:
                        col_real = j
                    elif "OBSERV" in c:
                        col_obs = j
                break
        if header_idx is None or col_real is None:
            continue
        for fila in tabla[header_idx + 1:]:
            item = limpia(fila[0]) if fila else ""
            if not es_item(item):
                continue
            desc = limpia(fila[1]) if len(fila) > 1 else ""
            teo = limpia(fila[col_teo]) if col_teo is not None and col_teo < len(fila) else ""
            real = limpia(fila[col_real]) if col_real < len(fila) else ""
            obs_c = limpia(fila[col_obs]) if col_obs and col_obs < len(fila) else ""
            if not real or real == "-":
                continue
            v_esp = teo if teo and teo not in ("-", "") else extraer_rango(desc)
            obs = obs_txt.get(item, obs_c if obs_c not in ("-", "") else "")
            filas.append(fila_base(
                enc, seccion, item, desc,
                ubicacion="CABINA",
                v_esp=v_esp, v_med=real, obs=obs
            ))
    return filas


def _lado_desde_contexto(tabla, header_idx: int, col_idx: int) -> str:
    for i in range(header_idx - 1, -1, -1):
        fila = tabla[i]
        for delta in range(min(5, len(fila))):
            for idx in [col_idx - delta, col_idx + delta]:
                if 0 <= idx < len(fila):
                    v = norm(limpia(fila[idx]))
                    if "DERECHO" in v:
                        return "D"
                    elif "IZQUIERDO" in v:
                        return "I"
                    elif "CENTRO" in v:
                        return "C"
    return ""


def parse_traccion(pagina: dict, enc: dict, seccion: str) -> list:
    """
    OK | OBSERVACION con columnas D/I.
    Extrae X y valores numéricos.
    """
    filas = []
    obs_txt = extraer_obs_texto(pagina["texto"])
    for tabla in pagina["tablas"]:
        if not tabla or len(tabla) < 3:
            continue
        header_idx = None
        cols_ok = []
        cols_obs = []
        for i, fila in enumerate(tabla[:7]):
            cn = norm(" ".join(limpia(c) for c in fila if c))
            if ("OK" in cn or "CONTROL" in cn):
                header_idx = i
                for j, celda in enumerate(fila):
                    c = norm(limpia(celda))
                    lado = _lado_desde_contexto(tabla, i, j)
                    if c in ("OK", "CONTROL"):
                        cols_ok.append((j, lado))
                    elif "OBSERVACI" in c:
                        cols_obs.append((j, lado))
                break
        if header_idx is None or not cols_ok:
            continue
        # Si no hay columna de observaciones, intentar detectarla en filas siguientes del header
        if not cols_obs:
            for i2 in range(header_idx + 1, min(header_idx + 3, len(tabla))):
                for j2, celda in enumerate(tabla[i2]):
                    if "OBSERVACI" in norm(limpia(celda)):
                        cols_obs.append((j2, ""))
                        break
                if cols_obs:
                    break
        for fila in tabla[header_idx + 1:]:
            item = limpia(fila[0]) if fila else ""
            if not es_item(item):
                continue
            desc = limpia(fila[1]) if len(fila) > 1 else ""
            v_esp = extraer_rango(desc)
            obs = obs_txt.get(item, "")
            for j, lado in cols_ok:
                if j >= len(fila):
                    continue
                valor = limpia(fila[j])
                vn = norm(valor)
                obs_val = obs
                if not obs_val:
                    for jo, lado_o in cols_obs:
                        if lado_o == lado and jo < len(fila):
                            v = limpia(fila[jo])
                            if v and v != "-":
                                obs_val = v
                            break
                if vn == "X":
                    filas.append(fila_base(
                        enc, seccion, item, desc, lado=lado,
                        ubicacion=seccion.replace("_", " "),
                        v_esp=v_esp if v_esp else "OK",
                        v_med="X", obs=obs_val
                    ))
                elif re.match(r"^[\d.,]+$", valor.replace(" ", "")):
                    filas.append(fila_base(
                        enc, seccion, item, desc, lado=lado,
                        ubicacion=seccion.replace("_", " "),
                        v_esp=v_esp, v_med=valor, obs=obs_val
                    ))
        # También capturar numéricos en columnas no detectadas como cols_ok
        cols_ok_idx = {j for j, _ in cols_ok}
        for j, celda in enumerate(fila):
            if j in cols_ok_idx or j < 2:
                continue
            valor = limpia(celda)
            if valor and re.match(r"^\d[\d.,\s]*(?:[a-zA-Z%°]+\.?)?$", valor.strip()):
                filas.append(fila_base(
                    enc, seccion, item, desc, lado="",
                    ubicacion=seccion.replace("_", " "),
                    v_esp=v_esp, v_med=valor, obs=obs
                ))
    return filas


def parse_frenos_mecanica_multirueda(pagina: dict, enc: dict, seccion: str) -> list:
    """
    Parser para sección 3 de PISE-024: tabla de 21 cols con header partido en 2 filas.
    Fila 0: ITEMS | TAREAS | _ | BOGIE1 x8 | BOGIE2 x8
    Fila 1: _ | _ | EJE1 x4 | EJE2 x4 | EJE3 x4 | EJE4 x4
    Fila 2: _ | _ | RUEDA1 | _ | _ | RUEDA(2 partido) | _ | RUEDA3 | ... (parcial)
    Fila 3: _ | _ | _ | _ | _ | 2 | _ | ...           (resto del número)
    Datos en cols pares: 2,4,7,9,11,13,16,19
    """
    filas = []
    obs_txt = extraer_obs_texto(pagina["texto"])
    for tabla in pagina["tablas"]:
        if not tabla or len(tabla) < 5:
            continue
        # Detectar tabla de frenos con >10 cols y header RUEDA en filas 2-3
        if len(tabla[0]) < 10:
            continue
        # Buscar filas de header (RUEDA partido en 2 filas)
        header_rueda_idx = None
        for i, fila in enumerate(tabla[:6]):
            cn = norm(" ".join(limpia(c) for c in fila if c))
            if re.search(r"RUEDA\s*\d+", cn):
                header_rueda_idx = i
                break
        if header_rueda_idx is None:
            continue

        # Construir mapa de columnas combinando filas header_rueda_idx y header_rueda_idx+1
        # para reconstruir nombres tipo "RUEDA 2" que pdfplumber parte en 2 filas
        fila_r1 = tabla[header_rueda_idx]
        fila_r2 = tabla[header_rueda_idx + 1] if header_rueda_idx + 1 < len(tabla) else []
        fila_eje = tabla[header_rueda_idx - 1] if header_rueda_idx > 0 else []
        fila_bogie = tabla[header_rueda_idx - 2] if header_rueda_idx > 1 else []

        col_map = {}  # col_idx → (bogie, rueda, lado, eje_ctx)
        for j, celda in enumerate(fila_r1):
            parte1 = limpia(celda)
            parte2 = limpia(fila_r2[j]) if j < len(fila_r2) else ""
            nombre = (parte1 + " " + parte2).strip()
            rm = re.search(r"RUEDA\s*(\d+)", norm(nombre))
            if not rm and parte2:
                rm = re.search(r"RUEDA\s*(\d+)", norm(parte1 + parte2))
            if rm:
                rueda = rm.group(1)
                lado_m = re.search(r"\(([ID])\)", nombre)
                lado = lado_m.group(1) if lado_m else ""
                # Buscar EJE y BOGIE en filas anteriores
                eje_ctx = ""
                for fi in [fila_eje]:
                    for delta in range(-2, 3):
                        idx = j + delta
                        if 0 <= idx < len(fi):
                            v = norm(limpia(fi[idx]))
                            if re.search(r"EJE\s*\d+", v):
                                eje_ctx = limpia(fi[idx])
                                break
                bogie = ""
                for fi in [fila_bogie]:
                    for delta in range(-4, 5):
                        idx = j + delta
                        if 0 <= idx < len(fi):
                            bm = re.search(r"BOGIE\s*(?:N[°º])?\s*(\d+)", norm(limpia(fi[idx])))
                            if bm:
                                bogie = bm.group(1)
                                break
                col_map[j] = (bogie, rueda, lado, eje_ctx)

        if not col_map:
            continue

        # Datos en filas desde header_rueda_idx + 2 (saltando las 2 filas de header)
        data_start = header_rueda_idx + 2
        # Si la fila siguiente al header aún no tiene ítem (puede ser la 2da parte del header)
        if data_start < len(tabla) and not es_item(limpia(tabla[data_start][0] if tabla[data_start] else "")):
            data_start += 1

        for fila in tabla[data_start:]:
            item = limpia(fila[0]) if fila else ""
            if not es_item(item):
                continue
            desc = limpia(fila[1]) if len(fila) > 1 else ""
            v_esp = extraer_rango(desc)
            obs = obs_txt.get(item, "")
            for j, (bogie, rueda, lado, eje_ctx) in col_map.items():
                if j >= len(fila):
                    continue
                valor = limpia(fila[j])
                if not valor or valor in ("-", "N/A", "OK"):
                    continue
                vn = norm(valor)
                if vn == "X" or re.match(r"^[\d.,/\s]+$", valor.replace(" ", "")):
                    filas.append(fila_base(
                        enc, seccion, item, desc,
                        bogie=bogie, rueda=rueda, lado=lado,
                        ubicacion=eje_ctx or f"BOGIE {bogie}",
                        v_esp=v_esp, v_med=valor, obs=obs
                    ))
    return filas
    for i in range(header_idx - 1, -1, -1):
        fila = tabla[i]
        for delta in range(min(5, len(fila))):
            for idx in [col_idx - delta, col_idx + delta]:
                if 0 <= idx < len(fila):
                    v = norm(limpia(fila[idx]))
                    if "DERECHO" in v:
                        return "D"
                    elif "IZQUIERDO" in v:
                        return "I"
                    elif "CENTRO" in v:
                        return "C"
    return ""


def parse_generico_x(pagina: dict, enc: dict, seccion: str) -> list:
    """Secciones 5-10: extrae solo las filas con X."""
    filas = []
    obs_txt = extraer_obs_texto(pagina["texto"])
    for tabla in pagina["tablas"]:
        if not tabla or len(tabla) < 3:
            continue
        header_idx = col_ctrl = col_obs = None
        for i, fila in enumerate(tabla[:5]):
            cn = norm(" ".join(limpia(c) for c in fila if c))
            if "CONTROL" in cn or "OBSERVACI" in cn:
                header_idx = i
                for j, celda in enumerate(fila):
                    c = norm(limpia(celda))
                    if "CONTROL" in c and col_ctrl is None:
                        col_ctrl = j
                    elif "OBSERVACI" in c and col_obs is None:
                        col_obs = j
                break
        if header_idx is None or col_ctrl is None:
            continue

        # Fix PISE-024: header 'CONTROL' en col 3 pero datos X/OK en col 2.
        # Verificar con primera fila de dato: si col_ctrl está vacío y col 2 tiene dato.
        primera = next(
            (f for f in tabla[header_idx + 1:] if es_item(limpia(f[0]) if f else "")),
            None
        )
        if primera is not None:
            dato_en_ctrl = col_ctrl < len(primera) and limpia(primera[col_ctrl]) not in ("", "-")
            dato_en_col2 = len(primera) > 2 and limpia(primera[2]) not in ("", "-")
            if not dato_en_ctrl and dato_en_col2:
                col_ctrl = 2

        for fila in tabla[header_idx + 1:]:
            item = limpia(fila[0]) if fila else ""
            if not es_item(item):
                continue
            desc = limpia(fila[1]) if len(fila) > 1 else ""
            ctrl = limpia(fila[col_ctrl]) if col_ctrl < len(fila) else ""
            obs_c = limpia(fila[col_obs]) if col_obs and col_obs < len(fila) else ""
            ctrl_norm = norm(ctrl)
            es_x = ctrl_norm == "X"
            es_num = bool(re.search(r"\d[\d\.,]*\s*(?:[a-zA-Z%\u00b0/\u00b7]+\.?)?$", ctrl.strip())
                          and re.match(r"^[\d\-]", ctrl.strip()))
            if not es_x and not es_num:
                continue
            v_esp = extraer_rango(desc)
            obs = obs_txt.get(item, obs_c if obs_c not in ("-", "") else "")
            filas.append(fila_base(
                enc, seccion, item, desc,
                ubicacion=seccion.replace("_", " "),
                v_esp=v_esp if v_esp else "OK",
                v_med=ctrl, obs=obs
            ))
    return filas


# ══════════════════════════════════════════════════════════
#  DISPATCHER
# ══════════════════════════════════════════════════════════

def procesar_pagina(pagina: dict, seccion: str, enc: dict) -> list:
    if seccion == "1_RUEDAS":
        return parse_ruedas(pagina, enc, seccion)
    elif seccion in ("2_BOGIES", "2A_BOGIE1", "2B_BOGIE2"):
        return parse_bogies(pagina, enc, seccion)
    elif seccion == "2C_MESA_CENTRAL":
        return parse_mesa_central(pagina, enc, seccion)
    elif seccion == "3A_FRENOS_MECANICA":
        # Intentar parser multi-rueda (PISE-024 con 21 cols) y caer a bogies si no produce
        filas = parse_frenos_mecanica_multirueda(pagina, enc, seccion)
        return filas if filas else parse_bogies(pagina, enc, seccion)
    elif seccion == "3B_FRENOS_NEUMATICA":
        return parse_frenos_neumatica(pagina, enc, seccion)
    elif seccion in ("4_TRACCION_CHOQUE", "4A_TRACCION_CABINA", "4B_TRACCION_CONTRARIA"):
        return parse_traccion(pagina, enc, seccion)
    # PISE-024: Sección 8 Furgón/Motogenerador → genérico (OK/X + col observaciones)
    elif seccion == "8_FURGON_MOTOGEN":
        return parse_generico_x(pagina, enc, seccion)
    else:
        return parse_generico_x(pagina, enc, seccion)


# ══════════════════════════════════════════════════════════
#  PROCESADOR COMPLETO DE UN PDF
# ══════════════════════════════════════════════════════════

def _detectar_seccion_tabla(tabla: list) -> str | None:
    """
    Intenta detectar la sección de una tabla individual.
    Estrategia 1: buscar título de sección en las primeras filas de la tabla.
    Estrategia 2 (fallback): inferir desde el número de ítem de la primera fila de datos.
    """
    if not tabla:
        return None
    # Estrategia 1: texto de las primeras 3 filas
    texto = " ".join(
        limpia(c)
        for fila in tabla[:3]
        for c in fila
        if c
    )
    sec = detectar_seccion(texto)
    if sec:
        return sec

    # Estrategia 2: buscar el primer ítem de la tabla e inferir la sección
    # desde su número principal (ej: "5.1" → sección 5, "7.3" → sección 7)
    _ITEM_A_SECCION = {
        "1":  "1_RUEDAS",
        "2":  "2_BOGIES",          # genérico; se refinará por contexto
        "3":  "3A_FRENOS_MECANICA",
        "4":  "4_TRACCION_CHOQUE",
        "5":  "5_BAJO_BASTIDOR",
        "6":  "6_INTERIOR_SALON",
        "7":  "7_ESTRUCTURA_EXT",
        "8":  "8_FURGON_MOTOGEN",
        "9":  "9_INTERIOR_CABINA",
        "10": "10_ATS",
    }
    for fila in tabla:
        if not fila:
            continue
        item = limpia(fila[0])
        if es_item(item):
            prefijo = item.split(".")[0]
            return _ITEM_A_SECCION.get(prefijo)

    return None


def procesar_pdf(pdf_path: str) -> list:
    filename = Path(pdf_path).name
    paginas = leer_pdf(pdf_path)
    enc = extraer_encabezado(paginas)
    enc["_filename"] = filename

    # Paso 1: leer D. Conclusiones → clasificaciones
    clasificaciones = extraer_clasificaciones(paginas)

    # Paso 2: extraer filas de secciones 1-10, tabla a tabla
    todas_filas = []
    seccion_actual = None

    for pagina in paginas:
        # Saltar páginas de conclusiones/fotos
        if re.search(r"[CD]\.\s*(CONCLUSIONES?|REGISTROS\s+FOTOGR)", pagina["texto"], re.IGNORECASE):
            continue

        # Actualizar sección desde el texto de la página completa (sticky)
        nueva_pag = detectar_seccion(pagina["texto"])
        if nueva_pag:
            seccion_actual = nueva_pag
        if seccion_actual is None:
            continue

        # Procesar tabla a tabla: cada tabla puede tener su propia sección
        tablas = pagina.get("tablas") or []
        for tabla in tablas:
            if not tabla or len(tabla) < 3:
                continue

            # Intentar detectar sección desde el encabezado de esta tabla concreta
            sec_tabla = _detectar_seccion_tabla(tabla)

            # Usar la sección de la tabla si es distinta a la de página,
            # o la de página como fallback (sticky)
            if sec_tabla and sec_tabla != seccion_actual:
                sec_usar = sec_tabla
            else:
                sec_usar = seccion_actual

            # Crear un pagina_virtual con solo esta tabla para los parsers existentes
            pagina_virtual = {
                "num": pagina["num"],
                "texto": pagina["texto"],
                "tablas": [tabla],
            }
            filas = procesar_pagina(pagina_virtual, sec_usar, enc)
            todas_filas.extend(filas)

    # Paso 2b: inferir Bogie desde el nombre de sección o desde la rueda cuando quedó vacío
    _SEC_BOGIE = {
        "2A_BOGIE1": "1", "2B_BOGIE2": "2",
    }
    for fila in todas_filas:
        if not fila.get("Bogie"):
            # Inferir desde sección
            b = _SEC_BOGIE.get(fila.get("Sección", ""), None)
            if b:
                fila["Bogie"] = b
            # En sección RUEDAS: ruedas 5-8 pertenecen a Bogie 2, 1-4 a Bogie 1
            elif fila.get("Sección") == "1_RUEDAS":
                try:
                    r = int(fila.get("Rueda", "0") or "0")
                    fila["Bogie"] = "2" if r >= 5 else "1"
                except ValueError:
                    pass

    # Paso 3: cruzar clasificación + ítem par desde Conclusiones
    for fila in todas_filas:
        item = fila.get("Ítem técnico", "")
        resultado = clasificaciones.get(item)
        if resultado:
            cat, item2, desc_conc = resultado
            fila["Clasificación"]  = cat
            fila["Ítem técnico 2"] = item2
            # Si la observación del pie estaba vacía, usar la de Conclusiones
            if not fila.get("Observación"):
                fila["Observación"] = desc_conc
        else:
            fila["Clasificación"]  = "SIN OBSERVACION"
            fila["Ítem técnico 2"] = ""

    return todas_filas


# ══════════════════════════════════════════════════════════
#  EXPORTACIÓN EXCEL
# ══════════════════════════════════════════════════════════

COLUMNAS = [
    "Archivo PDF", "Código PISE/Informe", "Vehículo", "Línea", "Lugar",
    "Fecha inspección", "Informe N°", "PISE código",
    "Sección", "Ítem técnico", "Ítem técnico 2", "Descripción",
    "Bogie", "Rueda", "Lado", "Ubicación",
    "Valor esperado", "Valor medido",
    "Observación", "Clasificación",
]

COLORES_CLASIF = {
    "CRITICA":         "FFCCCC",
    "NORMAL":          "FFF2CC",
    "REPARADA":        "CCFFCC",
    "RECHAZADA":       "FF9999",
    "SIN OBSERVACION": "FFFFFF",
}


def exportar_excel(filas: list) -> io.BytesIO:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    df = pd.DataFrame(filas) if filas else pd.DataFrame(columns=COLUMNAS)
    for col in COLUMNAS:
        if col not in df.columns:
            df[col] = ""
    df = df[COLUMNAS]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="PISE_Datos")

        df_obs = df[df["Clasificación"] != "SIN OBSERVACION"]
        if not df_obs.empty:
            df_obs.to_excel(writer, index=False, sheet_name="Observaciones")

        df_x = df[df["Valor medido"].apply(lambda v: norm(str(v)) == "X")]
        if not df_x.empty:
            df_x.to_excel(writer, index=False, sheet_name="Solo_X")

    buf.seek(0)
    wb = load_workbook(buf)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Header
        header_fill = PatternFill("solid", fgColor="2F4F8F")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        # Columna de clasificación
        col_clasif = None
        for j, cell in enumerate(ws[1], 1):
            if cell.value == "Clasificación":
                col_clasif = j
                break

        # Colorear filas
        if col_clasif:
            for row in ws.iter_rows(min_row=2):
                clasif_val = norm(str(row[col_clasif - 1].value or ""))
                color = COLORES_CLASIF.get(clasif_val, "FFFFFF")
                fill = PatternFill("solid", fgColor=color)
                for cell in row:
                    cell.fill = fill

        # Anchos de columna
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 55)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ══════════════════════════════════════════════════════════
#  INTERFAZ GRADIO
# ══════════════════════════════════════════════════════════

def procesar_archivos(archivos):
    if not archivos:
        return "⚠️ No se cargaron archivos.", None

    todas_filas = []
    lines = []

    for archivo in archivos:
        try:
            filas = procesar_pdf(archivo.name)
            todas_filas.extend(filas)
            n = len(filas)
            n_x = sum(1 for f in filas if norm(str(f.get("Valor medido", ""))) == "X")
            nc = sum(1 for f in filas if f.get("Clasificación") == "CRITICA")
            nn = sum(1 for f in filas if f.get("Clasificación") == "NORMAL")
            nr = sum(1 for f in filas if f.get("Clasificación") == "REPARADA")
            nrec = sum(1 for f in filas if f.get("Clasificación") == "RECHAZADA")
            lines.append(
                f"✅ **{Path(archivo.name).name}**\n"
                f"   {n} registros  |  {n_x} con X\n"
                f"   🔴 Críticas: {nc}  🟡 Normales: {nn}  🟢 Reparadas: {nr}  ⚫ Rechazadas: {nrec}"
            )
        except Exception as e:
            import traceback
            lines.append(f"❌ **{Path(archivo.name).name}**: {str(e)}\n```\n{traceback.format_exc()}\n```")

    if not todas_filas:
        return "\n\n".join(lines) + "\n\n⚠️ No se extrajeron datos.", None

    excel_buf = exportar_excel(todas_filas)
    out_path = "/tmp/PISE_Extraccion.xlsx"
    with open(out_path, "wb") as f:
        f.write(excel_buf.read())

    n_tot = len(todas_filas)
    n_obs = sum(1 for f in todas_filas if f.get("Clasificación") != "SIN OBSERVACION")
    texto = "\n\n".join(lines)
    texto += f"\n\n---\n📊 Total registros: **{n_tot}**  |  📋 Con observación: **{n_obs}**"
    return texto, out_path


CSS = ".gradio-container { max-width: 900px; margin: auto; }"

with gr.Blocks(title="PISE — Extractor", css=CSS) as demo:
    gr.Markdown("""
    # 🚂 PISE — Extractor de Informes de Inspección
    **Bureau Veritas Argentina**

    Cargá uno o más PDFs. El Excel tiene **3 hojas**:
    - **PISE_Datos** — todos los registros
    - **Observaciones** — ítems con observación (clasificados desde D. Conclusiones)
    - **Solo_X** — solo ítems con resultado X

    | Color | Clasificación |
    |---|---|
    | 🔴 Rojo claro | CRÍTICA |
    | 🟡 Amarillo | NORMAL (no crítica) |
    | 🟢 Verde | REPARADA durante la inspección |
    | ⚫ Rojo fuerte | RECHAZADA |
    | ⬜ Blanco | SIN OBSERVACIÓN |
    """)

    with gr.Row():
        with gr.Column():
            archivos_input = gr.File(
                label="📂 Cargar PDFs",
                file_types=[".pdf"],
                file_count="multiple"
            )
            btn = gr.Button("⚙️ Procesar PDFs", variant="primary")
        with gr.Column():
            resumen_out = gr.Markdown()
            excel_out = gr.File(label="📥 Descargar Excel")

    btn.click(fn=procesar_archivos, inputs=[archivos_input], outputs=[resumen_out, excel_out])

    gr.Markdown("""
    ---
    **Secciones soportadas:** 1 Ruedas · 2a/2b Bogies · 2c Mesa Central ·
    3a Frenos Mecánica · 3b Frenos Neumática · 4 Tracción y Choque ·
    5 Bajo Bastidor · 6 Interior · 7 Estructura · 8 Motores · 9 Cabina · 10 ATS

    **PISE compatibles:** PISE-SGBV-030 (Coches DMU) · PISE-SGBV-011 (Locomotoras GM)
    """)

if __name__ == "__main__":
    demo.launch()
